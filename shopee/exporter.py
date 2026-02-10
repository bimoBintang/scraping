"""
Shopee Data Exporter
Export wishlist and price history to various formats
"""

import csv
import json
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional

from .models import PriceHistory, ShopeeProduct, WishlistItem
from .wishlist import WishlistManager

logger = logging.getLogger(__name__)


class ShopeeExporter:
    """
    Export scraping results and price history
    
    Supported formats: CSV, JSON, Excel
    """
    
    def __init__(self, wishlist: WishlistManager):
        self.wishlist = wishlist
    
    def export_wishlist(self, filepath: str, format: str = "csv"):
        """Export wishlist items"""
        items = self.wishlist.list_all(active_only=False)
        
        if format == "csv":
            self._export_wishlist_csv(items, filepath)
        elif format == "json":
            self._export_wishlist_json(items, filepath)
        elif format == "excel":
            self._export_wishlist_excel(items, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        print(f"✅ Exported {len(items)} items to {filepath}")
    
    def export_price_history(
        self,
        item_id: int,
        filepath: str,
        format: str = "csv"
    ):
        """Export price history for a single product"""
        history = self.wishlist.get_price_history(item_id)
        
        if format == "csv":
            self._export_history_csv(history, filepath)
        elif format == "json":
            self._export_history_json(history, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        print(f"✅ Exported {len(history.snapshots)} snapshots to {filepath}")
    
    def export_all_history(self, output_dir: str, format: str = "csv"):
        """Export price history for all wishlist items"""
        os.makedirs(output_dir, exist_ok=True)
        
        items = self.wishlist.list_all(active_only=False)
        
        for item in items:
            ext = "csv" if format == "csv" else "json"
            filename = f"price_history_{item.item_id}.{ext}"
            filepath = os.path.join(output_dir, filename)
            
            self.export_price_history(item.item_id, filepath, format)
        
        print(f"✅ Exported history for {len(items)} items to {output_dir}/")
    
    def export_report(self, report: Dict, filepath: str):
        """Export price report to JSON"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Report exported to {filepath}")
    
    def export_search_results(
        self,
        products: List[ShopeeProduct],
        filepath: str,
        format: str = "csv"
    ):
        """Export search results"""
        if format == "csv":
            self._export_products_csv(products, filepath)
        elif format == "json":
            self._export_products_json(products, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        print(f"✅ Exported {len(products)} products to {filepath}")
    
    # ============ CSV Exports ============
    
    def _export_wishlist_csv(self, items: List[WishlistItem], filepath: str):
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Item ID", "Shop ID", "Nama Produk", "URL",
                "Harga Awal", "Harga Saat Ini", "Selisih (Rp)", "Selisih (%)",
                "Target Harga", "Status", "Ditambahkan", "Terakhir Dicek"
            ])
            
            for item in items:
                delta_rp = item.current_price - item.initial_price
                delta_pct = (delta_rp / item.initial_price * 100) if item.initial_price > 0 else 0
                status = "Aktif" if item.is_active else "Nonaktif"
                
                writer.writerow([
                    item.item_id, item.shop_id, item.product_name, item.product_url,
                    item.initial_price, item.current_price, round(delta_rp), round(delta_pct, 2),
                    item.target_price or "", status,
                    item.added_at[:19] if item.added_at else "",
                    item.last_checked[:19] if item.last_checked else "",
                ])
    
    def _export_history_csv(self, history: PriceHistory, filepath: str):
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Timestamp", "Harga", "Harga Original", "Diskon %",
                "Stok", "Terjual", "Selisih dari Awal (Rp)", "Selisih dari Awal (%)"
            ])
            
            for snap in history.snapshots:
                delta_rp = snap.price - history.initial_price
                delta_pct = (delta_rp / history.initial_price * 100) if history.initial_price > 0 else 0
                
                writer.writerow([
                    snap.timestamp[:19], snap.price, snap.price_original,
                    snap.discount_percent, snap.stock, snap.sold,
                    round(delta_rp), round(delta_pct, 2),
                ])
    
    def _export_products_csv(self, products: List[ShopeeProduct], filepath: str):
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Item ID", "Shop ID", "Nama", "Harga", "Harga Asli",
                "Diskon %", "Rating", "Terjual", "Stok", "Toko",
                "Lokasi", "URL"
            ])
            
            for p in products:
                writer.writerow([
                    p.item_id, p.shop_id, p.name, p.price, p.price_original,
                    p.discount_percent, p.rating, p.sold, p.stock,
                    p.shop_name, p.shop_location, p.product_url,
                ])
    
    # ============ JSON Exports ============
    
    def _export_wishlist_json(self, items: List[WishlistItem], filepath: str):
        data = [item.to_dict() for item in items]
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def _export_history_json(self, history: PriceHistory, filepath: str):
        data = history.to_dict()
        data["snapshots"] = [s.to_dict() for s in history.snapshots]
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def _export_products_json(self, products: List[ShopeeProduct], filepath: str):
        data = [p.to_dict() for p in products]
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    # ============ Excel Export ============
    
    def _export_wishlist_excel(self, items: List[WishlistItem], filepath: str):
        """Export to Excel (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
        except ImportError:
            logger.warning("openpyxl not installed. Install with: pip install openpyxl")
            # Fallback to CSV
            csv_path = filepath.replace('.xlsx', '.csv')
            self._export_wishlist_csv(items, csv_path)
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Wishlist Shopee"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="EE4D2D", end_color="EE4D2D", fill_type="solid")
        
        headers = [
            "Item ID", "Nama Produk", "Harga Awal", "Harga Kini",
            "Selisih (Rp)", "Selisih (%)", "Target", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, item in enumerate(items, 2):
            delta_rp = item.current_price - item.initial_price
            delta_pct = (delta_rp / item.initial_price * 100) if item.initial_price > 0 else 0
            
            ws.cell(row=row, column=1, value=item.item_id)
            ws.cell(row=row, column=2, value=item.product_name)
            ws.cell(row=row, column=3, value=item.initial_price)
            ws.cell(row=row, column=4, value=item.current_price)
            ws.cell(row=row, column=5, value=round(delta_rp))
            ws.cell(row=row, column=6, value=round(delta_pct, 2))
            ws.cell(row=row, column=7, value=item.target_price or "")
            ws.cell(row=row, column=8, value="Aktif" if item.is_active else "Nonaktif")
            
            # Color coding for price change
            delta_cell = ws.cell(row=row, column=5)
            if delta_rp < 0:
                delta_cell.font = Font(color="008000")  # Green = turun
            elif delta_rp > 0:
                delta_cell.font = Font(color="FF0000")  # Red = naik
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(filepath)
